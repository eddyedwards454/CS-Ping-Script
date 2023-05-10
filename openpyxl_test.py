import json
############https://realpython.com/openpyxl-excel-spreadsheets-python/#before-you-begin
from openpyxl import Workbook, load_workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

# workbook.save(filename="C:/SCRIPTS/SCRIPTS/Python Scripts/Learning Home Work/ACI Inventory/ACI_Node_Output_Nodes_Missingopenpyxl.xlsx")

# workbook = load_workbook(filename="C:/SCRIPTS/SCRIPTS/Python Scripts/Learning Home Work/ACI Inventory/ACI_Node_Output_Nodes_Missingopenpyxl.xlsx")
# print(workbook.sheetnames)
#
# sheet1 = workbook['ACI_Node']
# sheet2 = workbook['ACI_Node2']
# sheet2['E1'] = 'changed back by python again !!!'

#########Change a Cells value
# cell = sheet2['E1']
# cell.value = 'Changed by Python'


# sheet = workbook.active
# sheet['E1'] = "Wrtining"
workbook.save(filename="C:/SCRIPTS/SCRIPTS/Python Scripts/Learning Home Work/ACI Inventory/ACI_Node_Output_Nodes_Missingopenpyxl.xlsx")
# # print(sheet)
# def print_rows():
#     for row in sheet.iter_rows(values_only=True):
#         print(row)
# cellF = sheet['F1']
# cellG = sheet['F1']
# cellF.value = 'another test'
# cellG.value = 'another test G'
# workbook.save(filename="C:/SCRIPTS/SCRIPTS/Python Scripts/Learning Home Work/ACI Inventory/ACI_Node_Output_Nodes_Missingopenpyxl.xlsx")
# # print(sheet.title)
# #
# # print(sheet['A1'].value)
# #
# # print(sheet['A1:B2'])
# # print(sheet['A'])
# # print(sheet['A':'B'])
# # products = {}
# # for row in sheet.iter_rows(min_row=1,min_col=1,max_col=4,values_only=True):
# #     print(row)
# #     product_id = row[0]
# #     product = {
# #         'test': row[1],
# #         'test2': row[2]
# #
# #     }
# #     products[product_id] = product
# # print(json.dumps(products))
# # # for row in sheet.rows:
# # #     print(row.values)

